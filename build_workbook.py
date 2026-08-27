"""
build_workbook.py  --  master data -> Food_Recall_Tracker.xlsx
--------------------------------------------------------------
Reads data/recalls_master.csv and writes a multi-tab workbook:

    Read Me         orientation, sources, how to update, caveats
    Master Log      every recall, filterable (this feeds every other tab)
    By Region       recalls per Census region x class  (live SUMIFS)
    By Food Type    recalls per food type x class       (live COUNTIFS)
    By Cause        recalls per hazard category + agent->illness
    Trends          recalls per month + charts
    Policy Timeline regulatory events (with trend-confounder flags)
    Reference       hazard->illness table + class definitions

Aggregates are real formulas that point at the Master Log, so if you paste in
new rows or filter the log, the summaries and charts update. Region counts use
hidden 0/1 helper columns because a recall can span several regions.

Run:  python build_workbook.py
"""

import csv
import os
import re
import sys

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

import reference_data as ref

# Characters forbidden in the .xlsx (XML 1.0) format. openpyxl raises
# IllegalCharacterError on these, so we strip them from any incoming text.
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

MASTER = os.path.join("data", "recalls_master.csv")
OUT = "Food_Recall_Tracker.xlsx"

# palette
INK = "1F2933"
SLATE = "334155"
PAPER = "F8FAFC"
BAND = "EEF2F7"
CLASS1 = "F9D7D2"        # light red for Class I severity cue
CLASS1_TXT = "9B1C0F"
ACCENT = "0F5C6E"        # teal

FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE = Font(name=FONT, bold=True, color=INK, size=16)
SUB = Font(name=FONT, color=SLATE, size=10, italic=True)
BODY = Font(name=FONT, color=INK, size=10)
BOLD = Font(name=FONT, bold=True, color=INK, size=10)
HEAD_FILL = PatternFill("solid", fgColor=SLATE)
BAND_FILL = PatternFill("solid", fgColor=BAND)
thin = Side(style="thin", color="D5DCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


# Master Log display columns: (header, master_key, width, wrap)
COLS = [
    ("Agency", "agency", 11, False),
    ("Recall #", "recall_number", 16, False),
    ("Firm", "firm", 26, True),
    ("Product", "product_description", 34, True),
    ("Food Type", "food_type", 18, False),
    ("Reason", "reason", 34, True),
    ("Hazard Category", "hazard_category", 20, False),
    ("Agent (pathogen/allergen)", "agent", 22, False),
    ("Class", "classification", 22, False),
    ("Status", "status", 13, False),
    ("Distribution (raw)", "distribution_pattern", 26, True),
    ("States", "distribution_states", 20, True),
    ("Regions", "regions", 18, False),
    ("Nationwide", "nationwide", 11, False),
    ("Quantity (raw)", "quantity_raw", 24, True),
    ("Qty value", "quantity_value", 11, False),
    ("Qty unit", "quantity_unit", 10, False),
    ("Initiated", "date_initiated", 12, False),
    ("Reported", "date_reported", 12, False),
    ("Closed", "date_closed", 12, False),
    ("Days open", "days_open", 10, False),
    ("Year-Month", "ym", 11, False),
    ("Voluntary/Mandated", "voluntary_mandated", 20, True),
    ("Sample?", "is_sample", 9, False),
]
REGION_FLAGS = ["Northeast", "Midwest", "South", "West",
                "Territories", "Nationwide", "Unknown"]
CLASSES = ["Class I", "Class II", "Class III",
           "Public Health Alert (USDA)"]


def load_rows():
    if not os.path.exists(MASTER):
        sys.exit(f"Missing {MASTER}. Run:  python fetch_recalls.py --sample")
    with open(MASTER, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        for k, v in r.items():                 # strip Excel-illegal control chars
            if isinstance(v, str) and ILLEGAL_XML_RE.search(v):
                r[k] = ILLEGAL_XML_RE.sub(" ", v)
        yr, mo = r.get("year", ""), r.get("month", "")
        r["ym"] = f"{yr}-{mo}" if yr and mo else ""
    return rows


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def title_block(ws, title, subtitle, span=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.cell(1, 1, title).font = TITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.cell(2, 1, subtitle).font = SUB


# --------------------------------------------------------------------------- #
def sheet_master(wb, rows):
    ws = wb.create_sheet("Master Log")
    ws.sheet_properties.tabColor = SLATE
    headers = [c[0] for c in COLS] + [f"R_{r}" for r in REGION_FLAGS]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    first_flag = len(COLS) + 1

    for r in rows:
        line = []
        for _, key, _, _ in COLS:
            v = r.get(key, "")
            if key in ("quantity_value", "days_open") and v not in ("", None):
                try:
                    v = float(v) if key == "quantity_value" else int(float(v))
                except ValueError:
                    pass
            line.append(v)
        region_set = {x.strip() for x in (r.get("regions", "") or "").split(",")}
        line += [1 if fr in region_set else 0 for fr in REGION_FLAGS]
        ws.append(line)

    n = len(rows)
    last = n + 1
    # styling + Class I severity cue
    for row in range(2, last + 1):
        band = BAND_FILL if row % 2 == 0 else None
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY
            cell.border = BORDER
            _, key = (COLS[c - 1][0], COLS[c - 1][1]) if c <= len(COLS) else ("", "")
            cell.alignment = WRAP if (c <= len(COLS) and COLS[c - 1][3]) else TOP
            if band:
                cell.fill = band
        cls_cell = ws.cell(row=row, column=9)   # Class column
        if str(cls_cell.value).strip() == "Class I":
            cls_cell.fill = PatternFill("solid", fgColor=CLASS1)
            cls_cell.font = Font(name=FONT, bold=True, color=CLASS1_TXT, size=10)

    for i, (_, _, w, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for j in range(first_flag, first_flag + len(REGION_FLAGS)):
        ws.column_dimensions[get_column_letter(j)].hidden = True   # helper cols

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"
    return ws, last, first_flag


def col_letter_for(key):
    for i, (_, k, _, _) in enumerate(COLS, 1):
        if k == key:
            return get_column_letter(i)
    return None


def sheet_region(wb, last, first_flag):
    ws = wb.create_sheet("By Region")
    title_block(ws, "Recalls by region", "Region is derived from where each "
                "recall was distributed. A recall spanning several regions is "
                "counted in each. 'Nationwide' is its own bucket.", span=6)
    hdr = 4
    ws.append([]) if ws.max_row < hdr - 1 else None
    row = hdr
    ws.cell(row, 1, "Region").font = H
    for j, cls in enumerate(CLASSES, 2):
        ws.cell(row, j, cls).font = H
    ws.cell(row, len(CLASSES) + 2, "Total").font = H
    style_header(ws, row, len(CLASSES) + 2)

    cls_col = col_letter_for("classification")
    for i, region in enumerate(REGION_FLAGS):
        rr = hdr + 1 + i
        ws.cell(rr, 1, region).font = BOLD
        flag_col = get_column_letter(first_flag + i)
        for j, cls in enumerate(CLASSES, 2):
            # SUM of region flag where class matches
            f = (f"=SUMIFS('Master Log'!${flag_col}$2:${flag_col}${last},"
                 f"'Master Log'!${cls_col}$2:${cls_col}${last},$"
                 f"{get_column_letter(j)}${hdr})")
            ws.cell(rr, j, f).font = BODY
        tot_c = len(CLASSES) + 2
        ws.cell(rr, tot_c,
                f"=SUM(B{rr}:{get_column_letter(tot_c-1)}{rr})").font = BOLD
    # totals row
    tr = hdr + 1 + len(REGION_FLAGS)
    ws.cell(tr, 1, "All regions*").font = BOLD
    for j in range(2, len(CLASSES) + 3):
        cl = get_column_letter(j)
        ws.cell(tr, j, f"=SUM({cl}{hdr+1}:{cl}{tr-1})").font = BOLD
    ws.cell(tr + 1, 1, "*A multi-region recall is counted once per region, so "
            "this total exceeds the number of distinct recalls.").font = SUB
    ws.column_dimensions["A"].width = 16
    for j in range(2, len(CLASSES) + 3):
        ws.column_dimensions[get_column_letter(j)].width = 15
    _region_chart(ws, hdr, tr)
    return ws


def _region_chart(ws, hdr, tr):
    chart = BarChart()
    chart.type = "col"
    chart.title = "Recalls by region and class"
    chart.height, chart.width = 8, 16
    data = Reference(ws, min_col=2, max_col=1 + len(CLASSES),
                     min_row=hdr, max_row=tr - 1)
    cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=tr - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.title = "Recalls"
    ws.add_chart(chart, f"{get_column_letter(len(CLASSES)+4)}{hdr}")


def sheet_by_count(wb, tabname, title, subtitle, key, categories, last):
    ws = wb.create_sheet(tabname)
    title_block(ws, title, subtitle, span=6)
    hdr = 4
    ws.cell(hdr, 1, categories[0][0]).font = H
    for j, cls in enumerate(CLASSES, 2):
        ws.cell(hdr, j, cls).font = H
    ws.cell(hdr, len(CLASSES) + 2, "Total").font = H
    style_header(ws, hdr, len(CLASSES) + 2)

    key_col = col_letter_for(key)
    cls_col = col_letter_for("classification")
    cats = categories[1]
    for i, cat in enumerate(cats):
        rr = hdr + 1 + i
        ws.cell(rr, 1, cat).font = BOLD
        for j in range(2, len(CLASSES) + 2):
            f = (f"=COUNTIFS('Master Log'!${key_col}$2:${key_col}${last},$A{rr},"
                 f"'Master Log'!${cls_col}$2:${cls_col}${last},"
                 f"{get_column_letter(j)}${hdr})")
            ws.cell(rr, j, f).font = BODY
        tot_c = len(CLASSES) + 2
        ws.cell(rr, tot_c,
                f"=SUM(B{rr}:{get_column_letter(tot_c-1)}{rr})").font = BOLD
    tr = hdr + 1 + len(cats)
    ws.cell(tr, 1, "Total").font = BOLD
    for j in range(2, len(CLASSES) + 3):
        cl = get_column_letter(j)
        ws.cell(tr, j, f"=SUM({cl}{hdr+1}:{cl}{tr-1})").font = BOLD
    ws.column_dimensions["A"].width = 26
    for j in range(2, len(CLASSES) + 3):
        ws.column_dimensions[get_column_letter(j)].width = 15

    chart = BarChart(); chart.type = "bar"
    chart.title = title; chart.height, chart.width = 10, 16
    data = Reference(ws, min_col=len(CLASSES) + 2, max_col=len(CLASSES) + 2,
                     min_row=hdr, max_row=tr - 1)
    catref = Reference(ws, min_col=1, min_row=hdr + 1, max_row=tr - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(catref)
    chart.legend = None
    ws.add_chart(chart, f"{get_column_letter(len(CLASSES)+4)}{hdr}")
    return ws


def sheet_cause(wb, rows, last):
    # food-type & hazard tabs reuse sheet_by_count; here we add agent->illness
    ws = wb.create_sheet("By Cause")
    title_block(ws, "Recalls by cause & hazard agent",
                "Left: hazard category. Right: specific biological/allergen agent "
                "and what it can cause (from the Reference tab).", span=8)
    hazards = ["Biological (pathogen)", "Undeclared allergen", "Foreign material",
               "Chemical / contaminant", "Processing / production",
               "Labeling / quality", "Other / Unspecified", "Unspecified"]
    hdr = 4
    ws.cell(hdr, 1, "Hazard category").font = H
    ws.cell(hdr, 2, "Recalls").font = H
    ws.cell(hdr, 3, "Class I of those").font = H
    style_header(ws, hdr, 3)
    hz_col = col_letter_for("hazard_category")
    cls_col = col_letter_for("classification")
    for i, hz in enumerate(hazards):
        rr = hdr + 1 + i
        ws.cell(rr, 1, hz).font = BOLD
        ws.cell(rr, 2,
                f"=COUNTIF('Master Log'!${hz_col}$2:${hz_col}${last},$A{rr})").font = BODY
        ws.cell(rr, 3,
                f"=COUNTIFS('Master Log'!${hz_col}$2:${hz_col}${last},$A{rr},"
                f"'Master Log'!${cls_col}$2:${cls_col}${last},\"Class I\")").font = BODY
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15

    # agent -> count -> potential illness (INDEX/MATCH into Reference).
    # Pathogens are stored per-agent; allergen/foreign-material recalls carry the
    # specific agent (e.g. "Peanut") so those reference rows are counted by
    # hazard_category instead, to avoid undercounting.
    agents = [(r["agent"], r["category"]) for r in ref.HAZARD_ILLNESS]
    start = hdr
    ac1 = 5  # column E
    ws.cell(start, ac1, "Agent").font = H
    ws.cell(start, ac1 + 1, "Recalls").font = H
    ws.cell(start, ac1 + 2, "Potential illness").font = H
    ws.cell(start, ac1 + 3, "Higher-risk groups").font = H
    style_header(ws, start, ac1 + 3)
    agent_col = col_letter_for("agent")
    hz_col = col_letter_for("hazard_category")
    for i, (ag, cat) in enumerate(agents):
        rr = start + 1 + i
        e = get_column_letter(ac1)
        ws.cell(rr, ac1, ag).font = BOLD
        if cat == "Biological (pathogen)":
            cnt = f"=COUNTIF('Master Log'!${agent_col}$2:${agent_col}${last},${e}{rr})"
        else:  # count by category so specific allergens/foreign material roll up
            cnt = (f"=COUNTIF('Master Log'!${hz_col}$2:${hz_col}${last},"
                   f"'By Cause'!${e}{rr})")
        ws.cell(rr, ac1 + 1, cnt).font = BODY
        ws.cell(rr, ac1 + 2,
                f"=IFERROR(INDEX(Reference!$C:$C,MATCH(${e}{rr},Reference!$A:$A,0)),\"\")").font = BODY
        ws.cell(rr, ac1 + 3,
                f"=IFERROR(INDEX(Reference!$G:$G,MATCH(${e}{rr},Reference!$A:$A,0)),\"\")").font = BODY
    ws.column_dimensions[get_column_letter(ac1)].width = 24
    ws.column_dimensions[get_column_letter(ac1 + 1)].width = 10
    ws.column_dimensions[get_column_letter(ac1 + 2)].width = 40
    ws.column_dimensions[get_column_letter(ac1 + 3)].width = 40
    for rr in range(start + 1, start + 1 + len(agents)):
        for cc in (ac1 + 2, ac1 + 3):
            ws.cell(rr, cc).alignment = WRAP
    return ws


def sheet_trends(wb, rows, last):
    ws = wb.create_sheet("Trends")
    title_block(ws, "Recall trend over time",
                "Monthly counts. Read alongside the Policy Timeline: a dip can "
                "reflect reporting capacity (shutdown, staffing) rather than "
                "safer food.", span=6)
    # build the month axis from the data present
    yms = sorted({r["ym"] for r in rows if r.get("ym")})
    if not yms:
        yms = []
    hdr = 4
    ws.cell(hdr, 1, "Month").font = H
    ws.cell(hdr, 2, "All recalls").font = H
    ws.cell(hdr, 3, "Class I").font = H
    style_header(ws, hdr, 3)
    ym_col = col_letter_for("ym")
    cls_col = col_letter_for("classification")
    for i, ym in enumerate(yms):
        rr = hdr + 1 + i
        ws.cell(rr, 1, ym).font = BOLD
        ws.cell(rr, 2,
                f"=COUNTIF('Master Log'!${ym_col}$2:${ym_col}${last},$A{rr})").font = BODY
        ws.cell(rr, 3,
                f"=COUNTIFS('Master Log'!${ym_col}$2:${ym_col}${last},$A{rr},"
                f"'Master Log'!${cls_col}$2:${cls_col}${last},\"Class I\")").font = BODY
    tr = hdr + len(yms)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    if yms:
        chart = LineChart()
        chart.title = "Recalls per month"
        chart.height, chart.width = 9, 18
        data = Reference(ws, min_col=2, max_col=3, min_row=hdr, max_row=tr)
        cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=tr)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.title = "Recalls"
        ws.add_chart(chart, "E4")
    ws.cell(tr + 3, 1, "Note: counts reflect recalls reported in each month, "
            "not when illness occurred. A single event can span months.").font = SUB
    return ws


def sheet_policy(wb):
    ws = wb.create_sheet("Policy Timeline")
    title_block(ws, "Policy & regulatory timeline",
                "Rules and events that can change recall counts or response. "
                "'Confounder' items can shift reported numbers independent of "
                "actual food safety.", span=6)
    hdr = 4
    heads = ["Date", "Event", "Summary", "Effect on recalls",
             "Confidence", "Source"]
    for j, htext in enumerate(heads, 1):
        ws.cell(hdr, j, htext)
    style_header(ws, hdr, len(heads))
    widths = [12, 34, 46, 40, 22, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, e in enumerate(ref.POLICY_TIMELINE):
        rr = hdr + 1 + i
        vals = [e["date"], e["title"], e["summary"], e["effect_on_recalls"],
                e["confidence"], e["source"]]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, v)
            c.font = BODY
            c.alignment = WRAP
            c.border = BORDER
        if "CONFOUNDER" in e["effect_on_recalls"]:
            ws.cell(rr, 4).fill = PatternFill("solid", fgColor="FFF3C4")
    return ws


def sheet_reference(wb):
    ws = wb.create_sheet("Reference")
    title_block(ws, "Hazard -> potential illness (educational)",
                "General CDC/FDA/USDA-level information. Not medical advice. "
                "Used by the 'By Cause' tab to describe biological agents.", span=7)
    hdr = 4
    heads = ["Agent", "Category", "Potential illness", "Typical symptoms",
             "Usual onset", "Confidence", "Higher-risk groups", "Source"]
    # NOTE: column order here defines the INDEX/MATCH targets used in By Cause:
    #   A=Agent  C=Potential illness  G=Higher-risk groups
    order = ["agent", "category", "illness", "symptoms", "onset",
             "confidence", "higher_risk", "source"]
    for j, htext in enumerate(heads, 1):
        ws.cell(hdr, j, htext)
    style_header(ws, hdr, len(heads))
    widths = [24, 20, 26, 44, 22, 20, 40, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i, rowd in enumerate(ref.HAZARD_ILLNESS):
        rr = hdr + 1 + i
        for j, keyname in enumerate(order, 1):
            c = ws.cell(rr, j, rowd[keyname])
            c.font = BODY
            c.alignment = WRAP
            c.border = BORDER

    # class definitions below
    base = hdr + 2 + len(ref.HAZARD_ILLNESS)
    ws.cell(base, 1, "Recall classifications").font = TITLE
    for k, (name, desc) in enumerate(ref.CLASS_DEFINITIONS):
        rr = base + 1 + k
        ws.cell(rr, 1, name).font = BOLD
        c = ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
        cell = ws.cell(rr, 2, desc)
        cell.font = BODY
        cell.alignment = WRAP
    return ws


def sheet_readme(wb, rows):
    ws = wb.create_sheet("Read Me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    ws.cell(2, 2, "US Food Recall Tracker").font = Font(name=FONT, bold=True,
                                                        color=INK, size=20)
    n = len(rows)
    n_sample = sum(1 for r in rows if r.get("is_sample") == "Yes")
    lines = [
        ("", ""),
        (f"{n} recalls loaded" + (f"  ({n_sample} are SAMPLE rows)" if n_sample else ""), "sub"),
        ("", ""),
        ("What this is", "h"),
        ("A living log of US food recalls from both federal agencies, with "
         "views by region, food type, and cause, plus a trend view and a "
         "policy timeline. Rebuilt from data/recalls_master.csv each run.", "b"),
        ("", ""),
        ("Tabs", "h"),
        ("Master Log — every recall; filter it and the summaries follow.", "b"),
        ("By Region — recalls per Census region x class (derived from where "
         "each recall was distributed).", "b"),
        ("By Food Type / By Cause — counts by product category and by hazard, "
         "with what each biological agent can cause.", "b"),
        ("Trends — recalls per month, with charts.", "b"),
        ("Policy Timeline — rules/events; yellow = trend confounder.", "b"),
        ("Reference — hazard->illness table and class definitions.", "b"),
        ("", ""),
        ("Sources", "h"),
        ("FDA — openFDA Food Enforcement API (FDA-regulated foods).", "b"),
        ("USDA — FSIS Recall API (meat, poultry, egg).", "b"),
        ("CDC / FDA / USDA consumer pages for the hazard->illness reference.", "b"),
        ("", ""),
        ("Keep it updated", "h"),
        ("1.  python fetch_recalls.py --year 2026     (pulls new recalls, "
         "merges into the master)", "b"),
        ("2.  python build_workbook.py                (rebuilds this file)", "b"),
        ("3.  python build_dashboard.py               (rebuilds the web view)", "b"),
        ("", ""),
        ("Read the data honestly", "h"),
        ("• Region is derived from distribution text; a multi-region recall is "
         "counted in each region it reached.", "b"),
        ("• Quantities are not standardized across agencies (FDA free text vs "
         "USDA pounds); 'Qty value/unit' is a best-effort parse — trust "
         "'Quantity (raw)' when they differ.", "b"),
        ("• 'Days open' runs from initiation to closure, or to today if the "
         "recall is still ongoing.", "b"),
        ("• A change in recall counts is NOT the same as a change in food "
         "safety. Staffing cuts or a government shutdown can lower reported "
         "recalls (see Policy Timeline).", "b"),
        ("• The hazard->illness table is educational, not medical advice.", "b"),
    ]
    r = 3
    for text, kind in lines:
        cell = ws.cell(r, 2, text)
        if kind == "h":
            cell.font = Font(name=FONT, bold=True, color=ACCENT, size=12)
        elif kind == "sub":
            cell.font = Font(name=FONT, italic=True, color=SLATE, size=11)
        else:
            cell.font = BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    if n_sample:
        ws.cell(r + 1, 2, "⚠  This workbook currently contains SAMPLE rows so "
                "you can see the layout. Run fetch_recalls.py (without --sample) "
                "for live data.").font = Font(name=FONT, bold=True,
                                              color=CLASS1_TXT, size=11)
    return ws


def main():
    rows = load_rows()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_readme(wb, rows)
    _, last, first_flag = sheet_master(wb, rows)
    sheet_region(wb, last, first_flag)
    sheet_by_count(wb, "By Food Type", "Recalls by food type",
                   "Food type is inferred from each product description "
                   "(best-effort keyword match).", "food_type",
                   ("Food type", sorted({r["food_type"] for r in rows})), last)
    sheet_cause(wb, rows, last)
    sheet_trends(wb, rows, last)
    sheet_policy(wb)
    sheet_reference(wb)

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(rows)} recalls, {len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
